#!/usr/bin/env python3
"""
Extract the Stakeholder Engagement Matrix workbook into seed JSON for the app.

The three tabs number their stakeholders differently (MM's departments are
6.1.x in the Matrix but 6.2-6.5 in the Combined View), so stakeholders are
resolved by NAME through an explicit alias map rather than by row number.
Anything unmapped is reported, never guessed.
"""
import json, sys, re
from openpyxl import load_workbook

SRC = "/Users/mohamedhamza/Downloads/Stakeholder Engagement Matrix 20260819 (1).xlsx"
OUT = sys.argv[1] if len(sys.argv) > 1 else "engagement_seed_data.json"

# Workbook stakeholder name -> register sub_reference. GROUP = a parent header
# row that carries no rating of its own (KAHRAMAA, MM, MoI, ASHGHAL, WOQOD...).
GROUP = "__GROUP__"
ALIAS = {
    "Al Khor Park and Zoo": "AKPZ-S01",
    "Civil Aviation Authority (CAA)": "CAA-S01",
    "Gulf Organisation for Research and Development (GORD)": "GORD-S01",
    "Gulf Organization for Research and Development (GORD)": "GORD-S01",
    "KAHRAMAA": GROUP,
    "KAHRAMAA - Water": "KM-S01",
    "KAHRAMAA - Electricity": "KM-S02",
    "Ministry of Communications and IT (MCIT)": "MCIT-S01",
    "Ministry of Municipality (MM)": GROUP,
    "Ministry of Municipality (MM) - Urban Planning Department": "MM-S01",
    "MM - Urban Planning Department": "MM-S01",
    "MM - Live Stock (Animal Resources)": "MM-S02",
    "MM - Urban Planning - Live Stock (Animal Resources)": "MM-S02",
    "MM - Agriculture and Fisheries Affairs": "MM-S03",
    "MM - Urban Planning - Agriculture and Fisheries Affairs": "MM-S03",
    "MM - Cleaning Department": "MM-S04",
    "MM - Urban Planning - Cleaning Department": "MM-S04",
    "Ministry of Municipality (MM) - Public Parks": "MM-S05",
    "MM - Public Parks": "MM-S05",
    "Ministry of Environment and Climate Change (MoECC)": "MOECC-S01",
    "Ministry of Interior (MoI)": GROUP,
    "Qatar Civil Defence (QCDD)": "MOI-S01",
    "Qatar Civil Defense (QCDD)": "MOI-S01",
    "Security Services Department (SSD)": "MOI-S02",
    "General Directorate of Traffic (Traffic Police Dept.)": "MOI-S03",
    "MOI Telecom": "MOI-S04",
    "Establishments and Authorities Security Dep.": "MOI-S05",
    "Ministry of Public Health (MoPH)": "MOPH-S01",
    "Ministry of Transport (MOT)": "MOT-S01",
    "Ooredoo": "OOR-S01",
    "Public Works Authority (ASHGHAL)": GROUP,
    "ASHGHAL - Roads": "PWA-S01",
    "ASHGHAL - Drainage and Sewage": "PWA-S02",
    "ASHGHAL - TSE Network": "PWA-S03",
    "ASHGHAL - Beautification": "PWA-S04",
    "Qatar Armed Forces (QAF) - QESC": "QAF-S01",
    "Qatar Energy - Transmission and Distribution": "QE-S01",
    "Qatar National Broadband Network": GROUP,
    "Qatar National Broadband Network (QNBN)": GROUP,
    "Qatar National Broadband Network (QNBN) - General": "QNBN-S01",
    "(QNBN) - General": "QNBN-S01",
    "(QNBN) - UGN": "QNBN-S02",
    "Qatar Tourism": "QT-S01",
    "Vodafone": "VDF-S01",
    "Qatar Fuel Company (WOQOD)": GROUP,
    "(WOQOD) - GAS": "WOQOD-S01",
    "(WOQOD) - Diesel": "WOQOD-S02",
    "Qatar Museum Authority": "QMA-S01",
    "Ministry of Awqaf and Islamic Affairs": "AWQAF-S01",
    "Private Engineering Office (PEO)": "PEO-S01",
    "Mowasalat": "MWSL-S01",
    "Qatar Rail": "QRAIL-S01",
    "Marafeq": "MRFQ-S01",
    "Amiri Gurad": "AMIRI-S01",          # spelled this way in the workbook
    "Qatar cool": "QCOOL-S01",
}

# "Action By" spellings collapse onto one organisation each.
ORG_FIX = {
    "spd": "SPD", "spd/egis": None, "spd/ egis": None, "spd- authority": "SPD",
    "egis": "EGIS", "ecg": "ECG", "wa": "WA", "zoo solutions": "Zoo Solutions",
    "tc": "TC", "pwa": "PWA", "ashghal/ pwa": "PWA", "ashghal- roads": "PWA",
    "ashghal- tse": "PWA", "ashghal- foul": "PWA", "gord": "GORD",
    "mmup": "MM", "mmup-agri": "MM", "mmup-livestock": "MM",
    "mmup-planning": "MM", "mmup-p.parks": "MM", "mmup-cleaning": "MM",
    "mot": "MOT", "moe": "MoECC", "km- water": "KAHRAMAA", "km- elec": "KAHRAMAA",
    "qma": "QMA", "qta": "Qatar Tourism", "qnbn": "QNBN", "ooredoo": "Ooredoo",
    "vodafone": "Vodafone", "qatar rail": "Qatar Rail", "mowasalat": "Mowasalat",
    "awqaf": "AWQAF", "woqod- gas": "WOQOD", "woqod- diesel": "WOQOD",
    "peo": "PEO", "qaf": "QAF", "qe": "Qatar Energy", "moph": "MoPH",
    "ssd": "MOI-SSD", "qcdd": "QCDD", "mcit": "MCIT",
    "civil aviation": "CAA", "ak park & zoo": "Al Khor Park and Zoo",
}
INTERNAL = {"ECG", "SPD", "EGIS", "WA", "TC", "Zoo Solutions"}


def clean(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s if s and s not in ("—", "-") else None


def split_orgs(v):
    """'ECG / WA/ Zoo Solutions' -> ['ECG','WA','Zoo Solutions'] (SPD/EGIS -> both)."""
    s = clean(v)
    if not s:
        return []
    out = []
    for part in re.split(r"[/,&]| and ", s):
        p = part.strip()
        if not p:
            continue
        key = p.lower()
        if key in ("spd/egis", "spd/ egis"):
            out += ["SPD", "EGIS"]
            continue
        out.append(ORG_FIX.get(key, p))
    seen, res = set(), []
    for o in out:
        if o and o not in seen:
            seen.add(o)
            res.append(o)
    return res


def norm_milestone(v):
    s = clean(v)
    if not s:
        return None
    t = s.lower().replace("approval", "approval").strip()
    canon = {
        "upon approval of mp": "Upon Approval of MP",
        "upon approval of cd": "Upon Approval of CD",
        "upon approval of sd": "Upon Approval of SD",
        "upon submission of mp": "Upon Submission of MP",
        "during concept design stage": "During Concept Design Stage",
        "sd stage": "SD Stage",
        "in progress": None,          # a status, not a milestone
    }
    return canon.get(t, s)


def isodate(v):
    if v is None:
        return None
    try:
        return v.date().isoformat()
    except AttributeError:
        s = clean(v)
        return s if s and re.match(r"^\d{4}-\d{2}-\d{2}", s) else None


wb = load_workbook(SRC, data_only=True)
report = {"unmapped_names": [], "actions_without_date": 0, "actions_without_ref": 0,
          "actions_without_description": []}

# ---- ratings: Stakeholder Matrix (H/L) + Engagement Assessment (ladder) -----
ratings = {}


def put(subref, key, val):
    if not subref or subref == GROUP or val is None:
        return
    ratings.setdefault(subref, {})[key] = val


sm = wb["Stakeholder Matrix"]
for r in range(4, 49):
    nm = clean(sm.cell(r, 2).value)
    if not nm:
        continue
    ref = ALIAS.get(nm)
    if ref is None:
        report["unmapped_names"].append(f"Stakeholder Matrix r{r}: {nm}")
        continue
    put(ref, "influence", clean(sm.cell(r, 6).value))
    put(ref, "involvement", clean(sm.cell(r, 7).value))

ea = wb["Engagement Assessment"]
for r in range(5, 46):
    nm = clean(ea.cell(r, 2).value)
    if not nm:
        continue
    ref = ALIAS.get(nm)
    if ref is None:
        report["unmapped_names"].append(f"Engagement Assessment r{r}: {nm}")
        continue
    put(ref, "engagement_current", clean(ea.cell(r, 3).value))
    put(ref, "engagement_desired", clean(ea.cell(r, 4).value))
    put(ref, "gap_action_by", clean(ea.cell(r, 11).value))
    put(ref, "gap_remarks", clean(ea.cell(r, 12).value))

# ---- Combined View: communication priority + the action register ------------
cv = wb["Combined View"]
actions = []
cv_stakeholder_by_id = {}   # workbook Combined-View id -> sub_ref
current_ref = None

for r in range(4, 140):
    wid = cv.cell(r, 1).value
    if wid is None:
        continue
    wid = str(wid).strip()
    name = clean(cv.cell(r, 3).value)
    ref = ALIAS.get(name) if name else None

    if ref is not None:                       # a stakeholder row
        cv_stakeholder_by_id[wid] = ref
        if ref != GROUP:
            current_ref = ref
            put(ref, "communication_priority", clean(cv.cell(r, 17).value))
        continue

    # An action row. A handful carry no description at all — their substance
    # sits in the remarks column — so fall back to that rather than drop a
    # register entry, and report every one.
    desc = name
    remarks_raw = clean(cv.cell(r, 11).value)
    if not desc:
        has_content = any(clean(cv.cell(r, c).value) for c in (2, 5, 6, 7, 11))
        if not has_content:
            continue
        desc = remarks_raw or "(no description recorded in the source register)"
        report["actions_without_description"].append(f"r{r} id={wid}")
    # Prefer the id's own parent (6.5.21 -> 6.5); fall back to position.
    parent = wid.rsplit(".", 1)[0] if "." in wid else None
    target = cv_stakeholder_by_id.get(parent)
    if target in (None, GROUP):
        target = current_ref
    if not target:
        report["unmapped_names"].append(f"Combined View r{r}: action with no stakeholder")
        continue

    doc_ref = clean(cv.cell(r, 2).value)
    rec = isodate(cv.cell(r, 4).value)
    if not rec:
        report["actions_without_date"] += 1
    if not doc_ref:
        report["actions_without_ref"] += 1

    actions.append({
        "row": r,
        "wb_id": wid,
        "sub_ref": target,
        "description": desc,
        "doc_ref": doc_ref,
        "recorded_date": rec,
        "action_by": split_orgs(cv.cell(r, 5).value),
        "wb_status": clean(cv.cell(r, 6).value) or "Pending",
        "milestone": norm_milestone(cv.cell(r, 7).value),
        "remarks": None if desc == remarks_raw else remarks_raw,
    })

orgs = sorted({o for a in actions for o in a["action_by"]})
milestones = sorted({a["milestone"] for a in actions if a["milestone"]})

data = {
    "stakeholders": [dict(sub_ref=k, **v) for k, v in sorted(ratings.items())],
    "actions": actions,
    "orgs": [{"name": o, "is_internal": o in INTERNAL} for o in orgs],
    "milestones": milestones,
}
with open(OUT, "w") as f:
    json.dump(data, f, indent=1, ensure_ascii=False)

print(f"stakeholders with ratings : {len(data['stakeholders'])}")
print(f"actions                   : {len(actions)}")
print(f"organisations             : {len(orgs)}  -> {', '.join(orgs)}")
print(f"milestones                : {len(milestones)} -> {', '.join(milestones)}")
print(f"actions missing a date    : {report['actions_without_date']}")
print(f"actions missing a doc ref : {report['actions_without_ref']}")
from collections import Counter
print("workbook status split     :", dict(Counter(a['wb_status'] for a in actions)))
if report["actions_without_description"]:
    print("actions with no description (remark used instead):",
          len(report["actions_without_description"]),
          "->", ", ".join(report["actions_without_description"]))
if report["unmapped_names"]:
    print("\nUNMAPPED (would be guessed — fix the alias map):")
    for u in report["unmapped_names"]:
        print("   ", u)
else:
    print("\nevery stakeholder name resolved")
print(f"\nwritten -> {OUT}")
